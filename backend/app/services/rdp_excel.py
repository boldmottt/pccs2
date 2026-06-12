"""RDP-DB(rdp.db) ↔ 엑셀(xlsx) 변환 서비스.

- 양식/현재 데이터 엑셀 내보내기: 대량 수정·입력을 엑셀에서 하고 다시 업로드
- 엑셀 업로드: 고유키(project, pattern_code, plate, layer, batch_no) 기준 upsert
- PCCS2 → rdp_mixes 역변환: PCCS2에서 수정한 RDP 출신 샘플을 행으로 되돌림
"""

import io
import sqlite3
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook, load_workbook

from app.services.rdp_import import RDP_INK_COLUMNS

# (컬럼명, 타입, 설명) — rdp_mixes 스키마 그대로. 엑셀 헤더 행이 이 컬럼명을 사용한다.
RDP_EXCEL_COLUMNS: list[tuple[str, str, str]] = [
    ("date", "str", "작업일 (YYYY-MM-DD)"),
    ("project", "str", "차종/프로젝트 (필수)"),
    ("pattern_code", "str", "패턴 코드 (필수)"),
    ("plate", "str", "동판 번호 (필수)"),
    ("layer", "str", "도수 — 1도/2도 (필수)"),
    ("batch_no", "str", "배합 번호 (필수, 같은 날 같은 도수 내에서 구분)"),
    ("is_base", "int", "기준배합 여부 (1=기준, 0=파생)"),
    ("mt", "float", "MT 잉크 (g)"),
    ("bk", "float", "BK 잉크 (g)"),
    ("wh", "float", "WH 잉크 (g)"),
    ("ye", "float", "YE 잉크 (g)"),
    ("rd", "float", "RD 잉크 (g)"),
    ("cl", "float", "CL 투명 잉크 (g)"),
    ("ye_d", "float", "YE_D 잉크 (g)"),
    ("matting_agent_pct", "float", "소광제 비율 (%)"),
    ("matting_agent_g", "float", "소광제 (g)"),
    ("thinner_pct", "float", "신너 비율 (%)"),
    ("thinner_g", "float", "신너 (g)"),
    ("hardener_pct", "float", "경화제 비율 (%)"),
    ("hardener_g", "float", "경화제 (g)"),
    ("total_g", "float", "총량 (g)"),
    ("emboss_type", "str", "엠보 종류"),
    ("emboss_depth_um", "int", "엠보 깊이 (µm)"),
    ("coating_maker", "str", "도장 메이커"),
    ("coating_code", "str", "도장 코드"),
    ("coating_lot", "str", "도장 LOT"),
    ("pad_name", "str", "패드명"),
    ("pad_hardness", "str", "패드 경도"),
    ("result", "str", "결과 — ✅ / ⚠️ / ❌"),
    ("change_summary", "str", "변경 요약 (예: YE+3g)"),
    ("notes", "str", "비고"),
    ("source_file", "str", "출처 파일"),
    ("target_L", "float", "목표색 L* (SCI)"),
    ("target_a", "float", "목표색 a* (SCI)"),
    ("target_b", "float", "목표색 b* (SCI)"),
    ("target_sce_L", "float", "목표색 L* (SCE)"),
    ("target_sce_a", "float", "목표색 a* (SCE)"),
    ("target_sce_b", "float", "목표색 b* (SCE)"),
    ("measured_L", "float", "측정색 L* (SCI)"),
    ("measured_a", "float", "측정색 a* (SCI)"),
    ("measured_b", "float", "측정색 b* (SCI)"),
    ("measured_sce_L", "float", "측정색 L* (SCE)"),
    ("measured_sce_a", "float", "측정색 a* (SCE)"),
    ("measured_sce_b", "float", "측정색 b* (SCE)"),
    ("delta_e", "float", "ΔE (목표 대비, SCI)"),
    ("result_code", "str", "결과 코드"),
]

COLUMN_NAMES = [c[0] for c in RDP_EXCEL_COLUMNS]
REQUIRED_COLUMNS = ["date", "project", "pattern_code", "plate", "layer", "batch_no"]
KEY_COLUMNS = ["project", "pattern_code", "plate", "layer", "batch_no"]

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# rdp.db가 없거나 테이블이 없을 때 생성하는 스키마 (실제 rdp_mixes와 동일)
RDP_MIXES_SCHEMA = """
CREATE TABLE IF NOT EXISTS rdp_mixes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    date TEXT NOT NULL,
    project TEXT NOT NULL,
    pattern_code TEXT NOT NULL,
    plate TEXT NOT NULL,
    layer TEXT NOT NULL,
    batch_no TEXT NOT NULL,
    is_base INTEGER DEFAULT 0,
    mt REAL DEFAULT 0,
    bk REAL DEFAULT 0,
    wh REAL DEFAULT 0,
    ye REAL DEFAULT 0,
    rd REAL DEFAULT 0,
    cl REAL DEFAULT 0,
    ye_d REAL DEFAULT 0,
    matting_agent_pct REAL,
    matting_agent_g REAL,
    thinner_pct REAL,
    thinner_g REAL,
    hardener_pct REAL,
    hardener_g REAL,
    total_g REAL,
    emboss_type TEXT,
    emboss_depth_um INTEGER,
    coating_maker TEXT,
    coating_code TEXT,
    coating_lot TEXT,
    pad_name TEXT,
    pad_hardness TEXT,
    result TEXT,
    notes TEXT,
    change_summary TEXT,
    source_file TEXT,
    target_L REAL, target_a REAL, target_b REAL,
    target_sce_L REAL, target_sce_a REAL, target_sce_b REAL,
    measured_L REAL, measured_a REAL, measured_b REAL,
    measured_sce_L REAL, measured_sce_a REAL, measured_sce_b REAL,
    delta_e REAL, result_code TEXT,
    UNIQUE(project, pattern_code, plate, layer, batch_no)
)
"""

# 컬럼 타입 → SQLite 타입 (기존 rdp.db에 누락 컬럼을 ALTER로 추가할 때 사용)
_SQLITE_TYPES = {"str": "TEXT", "int": "INTEGER", "float": "REAL"}


def _ensure_columns(conn: sqlite3.Connection) -> list[str]:
    """기존 rdp_mixes에 없는 컬럼(예: SCE 측색값)을 ALTER로 보충한다."""
    existing = {r[1] for r in conn.execute("PRAGMA table_info(rdp_mixes)")}
    added = []
    for name, kind, _desc in RDP_EXCEL_COLUMNS:
        if name not in existing:
            conn.execute(
                f"ALTER TABLE rdp_mixes ADD COLUMN {name} {_SQLITE_TYPES[kind]}"
            )
            added.append(name)
    return added

EXAMPLE_ROW = {
    "date": "2026-05-11",
    "project": "NX5a",
    "pattern_code": "WB7",
    "plate": "26_027",
    "layer": "1도",
    "batch_no": "25",
    "is_base": 1,
    "mt": 2.3,
    "wh": 35.8,
    "ye": 25.0,
    "rd": 6.3,
    "thinner_pct": 30.0,
    "hardener_pct": 20.0,
    "result": "✅",
    "target_L": 73.0,
    "target_a": -2.1,
    "target_b": 5.2,
    "measured_L": 73.4,
    "measured_a": -1.9,
    "measured_b": 5.0,
    "delta_e": 0.55,
}


def _write_sheet(ws, rows: list[dict]) -> None:
    ws.append(COLUMN_NAMES)
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append([row.get(col) for col in COLUMN_NAMES])


def _add_description_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("컬럼 설명")
    ws.append(["컬럼", "설명"])
    for name, _kind, desc in RDP_EXCEL_COLUMNS:
        ws.append([name, desc])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50


def build_workbook(rows: list[dict], sheet_title: str = "rdp_mixes") -> bytes:
    """행 목록을 xlsx 바이트로 변환 (컬럼 설명 시트 포함)."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    _write_sheet(ws, rows)
    _add_description_sheet(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template() -> bytes:
    """빈 입력 양식 (예시 1행 포함)."""
    return build_workbook([EXAMPLE_ROW])


def _cell_to_value(name: str, kind: str, value):
    """엑셀 셀 값을 컬럼 타입으로 정규화. 빈 셀은 None."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
    if name == "date" and isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    if kind == "float":
        return float(value)
    if kind == "int":
        if isinstance(value, str):
            lowered = value.lower()
            if lowered in ("y", "yes", "true", "o"):
                return 1
            if lowered in ("n", "no", "false", "x"):
                return 0
        return int(float(value))
    return str(value)


def parse_workbook(content: bytes) -> tuple[list[dict], list[str]]:
    """업로드된 xlsx를 행 dict 목록으로 파싱. (rows, errors) 반환."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], ["빈 시트입니다"]

    col_index = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        h = str(h).strip()
        if h in COLUMN_NAMES:
            col_index[h] = i
    missing = [c for c in REQUIRED_COLUMNS if c not in col_index]
    if missing:
        return [], [f"필수 컬럼 누락: {', '.join(missing)} — 양식을 다운로드해 사용하세요"]

    kinds = {name: kind for name, kind, _ in RDP_EXCEL_COLUMNS}
    rows: list[dict] = []
    errors: list[str] = []
    for line_no, raw in enumerate(rows_iter, start=2):
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        row: dict = {}
        row_errors = []
        for name in COLUMN_NAMES:
            idx = col_index.get(name)
            value = raw[idx] if idx is not None and idx < len(raw) else None
            try:
                row[name] = _cell_to_value(name, kinds[name], value)
            except (TypeError, ValueError):
                row_errors.append(f"{line_no}행 {name}: 숫자가 아닙니다 ({value!r})")
        for req in REQUIRED_COLUMNS:
            if row.get(req) in (None, ""):
                row_errors.append(f"{line_no}행: {req} 값이 비어 있습니다")
        if row_errors:
            errors.extend(row_errors)
        else:
            rows.append(row)
    return rows, errors


def read_rdp_rows(db_path: str, project: Optional[str] = None) -> list[dict]:
    """rdp.db의 rdp_mixes 행을 dict 목록으로 읽는다 (보기/내보내기용)."""
    conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    try:
        tables = {
            r[0]
            for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
        }
        if "rdp_mixes" not in tables:
            raise ValueError("rdp_mixes 테이블이 없습니다 — 올바른 rdp.db 파일인지 확인하세요")
        sql = "SELECT * FROM rdp_mixes"
        params: tuple = ()
        if project:
            sql += " WHERE project = ?"
            params = (project,)
        sql += " ORDER BY date, id"
        result = []
        for row in conn.execute(sql, params):
            keys = row.keys()
            result.append({col: (row[col] if col in keys else None) for col in COLUMN_NAMES})
        return result
    finally:
        conn.close()


def upsert_rdp_rows(
    db_path: str,
    rows: list[dict],
    only_non_null: bool = False,
) -> dict:
    """행들을 rdp.db에 upsert. 고유키 일치 시 갱신, 없으면 추가.

    only_non_null=True 면 값이 있는 컬럼만 갱신한다 (PCCS2 역반영용 —
    PCCS2가 모르는 값으로 원본을 비우지 않도록).
    False(엑셀 업로드)면 빈 셀 = NULL 로 그대로 덮어쓴다.
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute(RDP_MIXES_SCHEMA)
        added_columns = _ensure_columns(conn)
        data_columns = [c for c in COLUMN_NAMES if c not in KEY_COLUMNS and c != "date"]
        inserted = 0
        updated = 0
        unchanged = 0
        for row in rows:
            key_values = [row[k] for k in KEY_COLUMNS]
            existing = conn.execute(
                "SELECT * FROM rdp_mixes WHERE project=? AND pattern_code=? AND plate=?"
                " AND layer=? AND batch_no=?",
                key_values,
            ).fetchone()
            if existing is None:
                cols = ["date"] + KEY_COLUMNS + data_columns
                placeholders = ",".join("?" for _ in cols)
                conn.execute(
                    f"INSERT INTO rdp_mixes ({','.join(cols)}) VALUES ({placeholders})",
                    [row.get(c) for c in cols],
                )
                inserted += 1
                continue

            set_cols = ["date"] + data_columns
            if only_non_null:
                set_cols = [c for c in set_cols if row.get(c) is not None]
            ek = existing.keys()
            changed = any(
                (existing[c] if c in ek else None) != row.get(c) for c in set_cols
            )
            if not changed:
                unchanged += 1
                continue
            assignments = ",".join(f"{c}=?" for c in set_cols)
            conn.execute(
                f"UPDATE rdp_mixes SET {assignments} WHERE project=? AND pattern_code=?"
                " AND plate=? AND layer=? AND batch_no=?",
                [row.get(c) for c in set_cols] + key_values,
            )
            updated += 1
        conn.commit()
        return {
            "inserted": inserted,
            "updated": updated,
            "unchanged": unchanged,
            "columns_added": added_columns,
        }
    finally:
        conn.close()


# PCCS2 success_flag → rdp result 이모지
FLAG_TO_RESULT = {"SUCCESS": "✅", "PENDING": "⚠️", "FAILED": "❌"}

# 잉크 이름 → rdp 컬럼 (MT → mt, YE_D → ye_d)
INK_NAME_TO_COLUMN = {name: col for col, (name, _cat) in RDP_INK_COLUMNS.items()}


def layer_to_rdp_row(
    layer: dict,
    work_date: Optional[str],
    ink_name_by_id: dict,
    emboss: Optional[tuple] = None,
) -> Optional[dict]:
    """PCCS2 샘플 레이어 하나를 rdp_mixes 행 dict로 역변환.

    rdp_key가 없는(RDP 출신이 아닌) 레이어는 None을 반환한다.
    """
    key = layer.get("rdp_key") or ""
    if not key.startswith("RDP:"):
        return None
    parts = key[4:].split("/")
    if len(parts) != 5:
        return None
    project, pattern_code, plate, layer_raw, batch_no = parts

    row: dict = {col: None for col in COLUMN_NAMES}
    row.update(
        {
            "date": work_date,
            "project": project,
            "pattern_code": pattern_code,
            "plate": plate,
            "layer": layer_raw,
            "batch_no": batch_no,
            "is_base": 1 if layer.get("is_base") else 0,
            "thinner_pct": layer.get("thinner_pct"),
            "thinner_g": layer.get("thinner_g"),
            "hardener_pct": layer.get("hardener_pct"),
            "hardener_g": layer.get("hardener_g"),
            "matting_agent_pct": layer.get("matting_agent_pct"),
            "matting_agent_g": layer.get("matting_agent_g"),
            "total_g": layer.get("total_g"),
            "coating_maker": layer.get("coating_maker"),
            "coating_code": layer.get("coating_code"),
            "coating_lot": layer.get("coating_lot"),
            "pad_name": layer.get("pad_name"),
            "pad_hardness": layer.get("pad_hardness"),
            "result": FLAG_TO_RESULT.get(layer.get("result") or ""),
            "change_summary": layer.get("change_summary"),
            "source_file": layer.get("source_file"),
            "delta_e": layer.get("delta_E_from_target"),
        }
    )
    for item in layer.get("ink_items") or []:
        name = item.get("ink_name") or ink_name_by_id.get(item.get("ink_id"))
        col = INK_NAME_TO_COLUMN.get(name or "")
        if col:
            row[col] = item.get("amount")
    color_columns = [
        ("target_color_sci", "target"),
        ("target_color_sce", "target_sce"),
        ("print_color_sci", "measured"),
        ("print_color_sce", "measured_sce"),
    ]
    for layer_field, prefix in color_columns:
        color = layer.get(layer_field)
        if color:
            for axis in ("L", "a", "b"):
                row[f"{prefix}_{axis}"] = color.get(axis)
    if emboss:
        row["emboss_type"], row["emboss_depth_um"] = emboss
    return row
