"""RDP-DB 엑셀 입출력·역반영 API 테스트."""

import io
import sqlite3

import pytest
from openpyxl import Workbook, load_workbook

from tests.test_rdp_import import RDP_SCHEMA, _upload, rdp_db_file  # noqa: F401


def _make_rdp_db(path, rows=()):
    conn = sqlite3.connect(path)
    conn.execute(RDP_SCHEMA)
    for row in rows:
        cols = ",".join(row.keys())
        placeholders = ",".join("?" for _ in row)
        conn.execute(f"INSERT INTO rdp_mixes ({cols}) VALUES ({placeholders})", list(row.values()))
    conn.commit()
    conn.close()
    return path


BASIC_ROW = {
    "date": "2026-05-11", "project": "NX5a", "pattern_code": "WB7",
    "plate": "26_027", "layer": "1도", "batch_no": "25", "is_base": 1,
    "mt": 2.3, "wh": 35.8, "ye": 25.0, "rd": 6.3,
    "thinner_pct": 30.0, "hardener_pct": 20.0, "result": "✅",
    "target_L": 73.0, "target_a": -2.1, "target_b": 5.2,
    "measured_L": 73.4, "measured_a": -1.9, "measured_b": 5.0, "delta_e": 0.55,
}


def _xlsx_bytes(rows_of_dicts):
    """헤더 + 행들로 xlsx 바이트 생성 (업로드 테스트용)."""
    from app.services.rdp_excel import COLUMN_NAMES
    wb = Workbook()
    ws = wb.active
    ws.append(COLUMN_NAMES)
    for row in rows_of_dicts:
        ws.append([row.get(c) for c in COLUMN_NAMES])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_template_download(api_client):
    resp = api_client.get("/api/rdp/excel/template")
    assert resp.status_code == 200
    assert "rdp_template.xlsx" in resp.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(resp.content))
    from app.services.rdp_excel import COLUMN_NAMES
    header = [c.value for c in next(wb.worksheets[0].iter_rows(max_row=1))]
    assert header == COLUMN_NAMES
    assert "컬럼 설명" in wb.sheetnames


def test_rows_endpoint_and_export(api_client, tmp_path):
    db_path = _make_rdp_db(tmp_path / "rdp.db", [BASIC_ROW])

    resp = api_client.get("/api/rdp/rows", params={"path": str(db_path)})
    assert resp.status_code == 200
    body = resp.json()
    assert body["total"] == 1
    assert body["projects"] == ["NX5a"]
    assert body["rows"][0]["batch_no"] == "25"
    assert body["rows"][0]["wh"] == 35.8

    resp = api_client.get("/api/rdp/excel/export", params={"path": str(db_path)})
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    rows = list(wb.worksheets[0].iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1


def test_rows_missing_file_404(api_client, tmp_path):
    resp = api_client.get("/api/rdp/rows", params={"path": str(tmp_path / "no.db")})
    assert resp.status_code == 404


def test_excel_upload_inserts_and_updates(api_client, tmp_path):
    db_path = tmp_path / "rdp.db"
    content = _xlsx_bytes([BASIC_ROW])
    resp = api_client.post(
        "/api/rdp/excel/upload",
        params={"path": str(db_path)},
        files={"file": ("bulk.xlsx", content, "application/octet-stream")},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["inserted"] == 1
    assert body["updated"] == 0
    assert body["errors"] == []

    # 같은 키, 값 수정 → 갱신 / 새 키 → 추가
    changed = dict(BASIC_ROW, ye=28.0, change_summary="YE+3g")
    new_row = dict(BASIC_ROW, layer="2도", batch_no="AH")
    content = _xlsx_bytes([changed, new_row])
    body = api_client.post(
        "/api/rdp/excel/upload",
        params={"path": str(db_path)},
        files={"file": ("bulk.xlsx", content, "application/octet-stream")},
    ).json()
    assert body["inserted"] == 1
    assert body["updated"] == 1

    conn = sqlite3.connect(db_path)
    ye, summary = conn.execute(
        "SELECT ye, change_summary FROM rdp_mixes WHERE layer='1도'"
    ).fetchone()
    conn.close()
    assert ye == 28.0
    assert summary == "YE+3g"


def test_upsert_rolls_back_on_failure(tmp_path):
    """배치 중간에 쓰기가 실패하면 그 배치 전체가 롤백돼 rdp.db가 일부만
    쓰인 상태로 남지 않아야 한다."""
    from app.services.rdp_excel import upsert_rdp_rows

    db = str(tmp_path / "rdp.db")
    good = {
        "date": "2026-05-11", "project": "P", "pattern_code": "C",
        "plate": "PL", "layer": "1도", "batch_no": "1", "mt": 1.0,
    }
    upsert_rdp_rows(db, [good])  # 스키마 생성 + 기준 행 1개

    # 같은 배치: 새 정상행 + date=None(NOT NULL 위반). 전체가 롤백돼야 한다.
    new_ok = dict(good, batch_no="2")
    bad = dict(good, batch_no="3", date=None)
    with pytest.raises(ValueError):
        upsert_rdp_rows(db, [new_ok, bad])

    conn = sqlite3.connect(db)
    count = conn.execute("SELECT COUNT(*) FROM rdp_mixes").fetchone()[0]
    conn.close()
    assert count == 1  # 기준 행만 남고 실패 배치(new_ok 포함)는 통째로 롤백


def test_excel_upload_validates_required(api_client, tmp_path):
    bad = dict(BASIC_ROW)
    bad.pop("project")
    content = _xlsx_bytes([bad])
    resp = api_client.post(
        "/api/rdp/excel/upload",
        params={"path": str(tmp_path / "rdp.db")},
        files={"file": ("bulk.xlsx", content, "application/octet-stream")},
    )
    assert resp.status_code == 400
    assert "project" in resp.json()["detail"]


def test_excel_upload_rejects_non_xlsx(api_client, tmp_path):
    resp = api_client.post(
        "/api/rdp/excel/upload",
        params={"path": str(tmp_path / "rdp.db")},
        files={"file": ("bad.xlsx", b"not an xlsx", "application/octet-stream")},
    )
    assert resp.status_code == 400


def test_export_pccs2_roundtrip(api_client, rdp_db_file):
    """가져오기 → PCCS2 형식 내보내기가 rdp_mixes 행으로 복원되는지 확인."""
    _upload(api_client, rdp_db_file)

    resp = api_client.get("/api/rdp/excel/export-pccs2")
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    from app.services.rdp_excel import COLUMN_NAMES
    ws = wb.worksheets[0]
    rows = [
        dict(zip(COLUMN_NAMES, values))
        for values in ws.iter_rows(min_row=2, values_only=True)
    ]
    assert len(rows) == 4  # 가져온 4행이 모두 복원

    base = next(r for r in rows if r["layer"] == "1도" and r["batch_no"] == "25")
    assert base["project"] == "NX5a"
    assert base["date"] == "2026-05-11"
    assert base["is_base"] == 1
    assert base["mt"] == 2.3
    assert base["wh"] == 35.8
    assert base["result"] == "✅"
    assert base["target_L"] == 73.0
    assert base["target_sce_L"] == 72.0
    assert base["measured_L"] == 73.4
    assert base["measured_sce_L"] == 72.5
    assert base["delta_e"] == 0.55
    assert base["emboss_type"] == "H-type"
    assert base["emboss_depth_um"] == 180


def test_sync_back_writes_pccs2_edits_to_rdp_db(api_client, rdp_db_file):
    """PCCS2에서 샘플을 수정하면 sync-back으로 rdp.db에 반영된다."""
    _upload(api_client, rdp_db_file)

    # PCCS2에서 기준배합 샘플의 YE 양 수정
    samples = api_client.get("/api/samples/").json()
    target = next(
        s for s in samples
        if any(ly.get("rdp_key") == "RDP:NX5a/WB7/26_027/1도/25" for ly in s["layers"])
    )
    layers = target["layers"]
    for ly in layers:
        for item in ly["ink_items"]:
            if item.get("ink_name") == "YE":
                item["amount"] = 27.5
    resp = api_client.put(f"/api/samples/{target['sample_id']}", json={"layers": layers})
    assert resp.status_code == 200
    # 수정 후에도 RDP 메타데이터가 유실되지 않아야 함
    assert resp.json()["layers"][0]["rdp_key"] == "RDP:NX5a/WB7/26_027/1도/25"

    body = api_client.post("/api/rdp/sync-back", json={"path": str(rdp_db_file)}).json()
    assert body["total_rows"] == 4
    assert body["updated"] >= 1

    conn = sqlite3.connect(rdp_db_file)
    ye = conn.execute(
        "SELECT ye FROM rdp_mixes WHERE layer='1도' AND batch_no='25'"
    ).fetchone()[0]
    conn.close()
    assert ye == 27.5


def test_upload_adds_missing_columns_to_old_db(api_client, tmp_path):
    """SCE 컬럼이 없는 구버전 rdp.db에 업로드하면 컬럼이 자동 추가된다."""
    db_path = tmp_path / "old.db"
    conn = sqlite3.connect(db_path)
    conn.execute(
        """CREATE TABLE rdp_mixes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL, project TEXT NOT NULL,
            pattern_code TEXT NOT NULL, plate TEXT NOT NULL,
            layer TEXT NOT NULL, batch_no TEXT NOT NULL,
            mt REAL, wh REAL, ye REAL, rd REAL,
            target_L REAL, target_a REAL, target_b REAL,
            UNIQUE(project, pattern_code, plate, layer, batch_no)
        )"""
    )
    conn.commit()
    conn.close()

    row = dict(BASIC_ROW, measured_sce_L=72.5, measured_sce_a=-1.8, measured_sce_b=4.8)
    body = api_client.post(
        "/api/rdp/excel/upload",
        params={"path": str(db_path)},
        files={"file": ("bulk.xlsx", _xlsx_bytes([row]), "application/octet-stream")},
    ).json()
    assert body["inserted"] == 1
    assert "measured_sce_L" in body["columns_added"]

    conn = sqlite3.connect(db_path)
    sce_l = conn.execute("SELECT measured_sce_L FROM rdp_mixes").fetchone()[0]
    conn.close()
    assert sce_l == 72.5


def test_custom_ink_full_flow(api_client, tmp_path):
    """PCCS2에 잉크 등록 → 양식에 컬럼 생성 → 업로드 → 가져오기 → 역반영."""
    # 1) PCCS2 잉크 마스터에 GR(녹색) 등록
    resp = api_client.post("/api/inks/", json={"ink_name": "GR", "ink_category": "COLOR"})
    assert resp.status_code == 201

    # 2) 양식 다운로드 → gr 컬럼이 ye_d 뒤에 추가됨
    resp = api_client.get("/api/rdp/excel/template")
    wb = load_workbook(io.BytesIO(resp.content))
    header = [c.value for c in next(wb.worksheets[0].iter_rows(max_row=1))]
    assert "gr" in header
    assert header.index("gr") == header.index("ye_d") + 1

    # 3) gr 값을 채워 업로드 → rdp.db에 gr 컬럼 자동 추가
    db_path = tmp_path / "rdp.db"
    row = dict(BASIC_ROW, gr=4.2)
    body = api_client.post(
        "/api/rdp/excel/upload",
        params={"path": str(db_path)},
        files={"file": ("bulk.xlsx", _xlsx_bytes_with_columns([row]), "application/octet-stream")},
    ).json()
    assert body["inserted"] == 1
    assert "gr" in body["columns_added"]

    # 4) 가져오기 → GR 잉크가 레이어 ink_items에 포함
    resp = api_client.post("/api/import/rdp/local", json={"path": str(db_path)})
    assert resp.json()["samples_created"] == 1
    samples = api_client.get("/api/samples/").json()
    _s, layer = _find_layer_any(samples)
    gr_item = next(i for i in layer["ink_items"] if i["ink_name"] == "GR")
    assert gr_item["amount"] == 4.2

    # 5) 역반영 → gr 컬럼으로 되돌아감
    api_client.post("/api/rdp/sync-back", json={"path": str(db_path)})
    conn = sqlite3.connect(db_path)
    gr = conn.execute("SELECT gr FROM rdp_mixes").fetchone()[0]
    conn.close()
    assert gr == 4.2


def _xlsx_bytes_with_columns(rows_of_dicts):
    """행 dict들의 모든 키를 헤더로 사용해 xlsx 생성 (추가 잉크 컬럼 포함)."""
    columns = list(dict.fromkeys(k for row in rows_of_dicts for k in row))
    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    for row in rows_of_dicts:
        ws.append([row.get(c) for c in columns])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _find_layer_any(samples):
    for s in samples:
        for ly in s["layers"] or []:
            if ly.get("rdp_key"):
                return s, ly
    raise AssertionError("no RDP layer found")


def test_favorite_blend_ink_becomes_template_column(api_client):
    """'기본잉크처럼 사용' 체크된 배합 잉크만 양식 컬럼에 포함된다."""
    # 배합 잉크 2개 생성: GRAY7(즐겨찾기), GRAY9(일반)
    for name in ("GRAY7", "GRAY9"):
        resp = api_client.post(f"/api/inks/blend-{name}/register-blend", json={
            "ink_name": name,
            "ink_category": "COLOR",
            "blend_recipe": {"ink_items": [{"ink_id": "x", "amount": 100}]},
        })
        assert resp.status_code == 200
        if name == "GRAY7":
            ink_id = resp.json()["ink_id"]
            upd = api_client.put(f"/api/inks/{ink_id}", json={"is_favorite": True})
            assert upd.status_code == 200
            assert upd.json()["is_favorite"] is True

    resp = api_client.get("/api/rdp/excel/template")
    wb = load_workbook(io.BytesIO(resp.content))
    header = [c.value for c in next(wb.worksheets[0].iter_rows(max_row=1))]
    assert "gray7" in header     # 즐겨찾기 배합 → 포함
    assert "gray9" not in header  # 일반 배합 → 제외


def test_sync_back_missing_file_404(api_client, tmp_path):
    resp = api_client.post("/api/rdp/sync-back", json={"path": str(tmp_path / "no.db")})
    assert resp.status_code == 404
